import os
import sqlite3
import threading
import time
from pathlib import Path

os.environ.setdefault("DATABASE_URL", "sqlite:///./test_persistence.db")

import requests
import uvicorn
from openpyxl import load_workbook

import psycopg2
from app.main import app


def _is_postgres(url: str) -> bool:
    return url.startswith("postgres://") or url.startswith("postgresql://")


def test_full_flow():
    # Ruta real de pruebas provista por el usuario
    excel_path = str((Path(__file__).parent / ".." / ".." / "pruebas" / "SAN_MARTIN_2025.xlsx").resolve())
    delegacion_path = str((Path(__file__).parent / ".." / ".." / "DELEGACION.xlsx").resolve())
    # Verificar que existan archivos de entrada
    assert os.path.exists(excel_path), f"No se encontró el Excel de pruebas: {excel_path}"
    assert os.path.exists(delegacion_path), f"No se encontró el Excel de ejemplo: {delegacion_path}"
    try:
        # Verificación previa: columnas A-Q en el Excel de entrada
        wb_in = load_workbook(excel_path)
        ws_in = wb_in.active
        assert ws_in.max_column >= 17, "El Excel de entrada debe contener al menos columnas A-Q"

        # Levantar servidor Uvicorn en hilo
        config = uvicorn.Config(app, host="127.0.0.1", port=8001, log_level="warning")
        server = uvicorn.Server(config)
        t = threading.Thread(target=server.run, daemon=True)
        t.start()
        # esperar a que levante
        time.sleep(0.8)

        # 1) Preparar hoja
        resp = requests.post("http://127.0.0.1:8001/sheet/prepare", json={"file_path": excel_path})
        assert resp.status_code == 200, resp.text
        data = resp.json()
        document_id = data["document_id"]
        assert data["rows_imported"] > 0

        # 1.b) Verificar que se hayan creado tablas en PostgreSQL si se usa Postgres
        db_url = os.getenv("DATABASE_URL", "")
        if _is_postgres(db_url):
            conn = psycopg2.connect(db_url)
            try:
                cur = conn.cursor()
                cur.execute("SELECT to_regclass('public.raw_incidents');")
                assert cur.fetchone()[0] is not None, "Tabla raw_incidents no existe en PostgreSQL"
                cur.execute("SELECT to_regclass('public.classified_incidents');")
                assert cur.fetchone()[0] is not None, "Tabla classified_incidents no existe en PostgreSQL"
            finally:
                conn.close()

        # 2) Obtener lote 1 (2 filas)
        resp = requests.get(f"http://127.0.0.1:8001/data/chunk/{document_id}?limit=2")
        assert resp.status_code == 200, resp.text
        chunk1 = resp.json()
        assert len(chunk1["items"]) == 2

        # 3) Guardar clasificación lote 1
        items_payload = []
        for it in chunk1["items"]:
            items_payload.append(
                {
                    "raw_incident_id": it["id"],
                    "col_r": "class_r",
                    "col_s": "class_s",
                    "col_t": "class_t",
                    "col_u": "class_u",
                    "col_v": "class_v",
                    "col_w": "class_w",
                    "col_x": "class_x",
                    "col_y": "class_y",
                    "col_z": "class_z",
                    "col_aa": "class_aa",
                    "col_ab": "class_ab",
                }
            )
        resp = requests.post(
            "http://127.0.0.1:8001/data/save_classified_chunk",
            json={"document_id": document_id, "items": items_payload},
        )
        assert resp.status_code == 200, resp.text
        assert resp.json()["saved"] == len(items_payload)

        # 4) Obtener lote 2 (resto)
        resp = requests.get(f"http://127.0.0.1:8001/data/chunk/{document_id}?limit=10")
        assert resp.status_code == 200, resp.text
        chunk2 = resp.json()
        assert len(chunk2["items"]) >= 0

        # 5) Guardar clasificación lote 2
        items_payload = []
        for it in chunk2["items"]:
            items_payload.append(
                {
                    "raw_incident_id": it["id"],
                    "col_r": "class_r",
                    "col_s": "class_s",
                    "col_t": "class_t",
                    "col_u": "class_u",
                    "col_v": "class_v",
                    "col_w": "class_w",
                    "col_x": "class_x",
                    "col_y": "class_y",
                    "col_z": "class_z",
                    "col_aa": "class_aa",
                    "col_ab": "class_ab",
                }
            )
        resp = requests.post(
            "http://127.0.0.1:8001/data/save_classified_chunk",
            json={"document_id": document_id, "items": items_payload},
        )
        assert resp.status_code == 200, resp.text
        assert resp.json()["saved"] == len(items_payload)

        # 6) Generar Excel final
        resp = requests.post(f"http://127.0.0.1:8001/sheet/generate_final/{document_id}")
        assert resp.status_code == 200, resp.text
        gen = resp.json()
        out_path = gen["file_path"]
        assert os.path.exists(out_path)

        # Validar formato: encabezado + filas, 28 columnas y color en R..AB
        wb2 = load_workbook(out_path)
        ws2 = wb2.active
        assert ws2.max_column == 28

        # Encabezados R..AB coloreados
        for col in range(18, 29):  # 18..28
            cell = ws2.cell(row=1, column=col)
            assert cell.fill.start_color.rgb in ("FFB2A1C7", "00FFB2A1C7", "B2A1C7")

        # Celdas de datos R..AB coloreadas también
        for row in range(2, ws2.max_row + 1):
            for col in range(18, 29):
                cell = ws2.cell(row=row, column=col)
                assert cell.fill.start_color.rgb in ("FFB2A1C7", "00FFB2A1C7", "B2A1C7")

        # Comparación básica con plantilla DELEGACION.xlsx en cuanto a color en R..AB
        wb_ref = load_workbook(delegacion_path)
        ws_ref = wb_ref.active
        # Color de referencia en celda R1
        ref_color = ws_ref.cell(row=1, column=18).fill.start_color.rgb
        out_color = ws2.cell(row=1, column=18).fill.start_color.rgb
        assert out_color in (ref_color, "FFB2A1C7", "00FFB2A1C7", "B2A1C7")
    finally:
        # Limpieza
        try:
            os.remove(f"final_{document_id}.xlsx")
        except Exception:
            pass


if __name__ == "__main__":
    test_full_flow()
    print("OK - test_full_flow completado")

