import logging
import os
import uuid
from typing import List
from pathlib import Path

from fastapi import FastAPI, HTTPException, Path, Query, UploadFile, File
from fastapi.responses import FileResponse

from .database import (
    fetch_unclassified_chunk,
    init_db,
    insert_classified_items,
    insert_raw_incident,
    fetch_raws,
    fetch_classified_map,
)
from .models import (
    ChunkResponse,
    GenerateFinalResponse,
    PrepareRequest,
    PrepareResponse,
    SaveClassifiedChunkRequest,
    SaveClassifiedChunkResponse,
)

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill


# Configuración de logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s - %(message)s",
)
logger = logging.getLogger("persistence_service")


app = FastAPI(title="Persistence Service - Proyecto Sentinel")

# Directorios de datos
from pathlib import Path as PathLib
DATA_DIR = PathLib("./data")
UPLOADS_DIR = DATA_DIR / "uploads"
FINAL_DIR = DATA_DIR / "final"


def ensure_directories():
    """Crea los directorios necesarios para uploads y archivos finales"""
    UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
    FINAL_DIR.mkdir(parents=True, exist_ok=True)
    
    logger.info("Directorios de datos creados/verificados")


@app.on_event("startup")
def on_startup():
    init_db()
    ensure_directories()
    logger.info("Base de datos inicializada y directorios creados")


@app.get("/health")
def health():
    return {"status": "ok"}


@app.post("/sheet/prepare", response_model=PrepareResponse)
def prepare_sheet(payload: PrepareRequest):
    """
    Recibe la ruta de un Excel, valida que tenga al menos columnas A-Q y persiste filas en BD.
    Devuelve un document_id para identificar el dataset.
    """
    try:
        file_path = payload.file_path
        if not os.path.exists(file_path):
            logger.error("Archivo no encontrado: %s", file_path)
            raise HTTPException(status_code=400, detail="El archivo no existe")

        wb = load_workbook(file_path)
        ws = wb.active

        # Validación: mínimo 17 columnas (A..Q)
        min_required_cols = 17
        if ws.max_column < min_required_cols:
            logger.error("El Excel no posee las columnas A-Q requeridas (encontradas: %s)", ws.max_column)
            raise HTTPException(status_code=400, detail="El Excel no posee las columnas A-Q requeridas")

        document_id = str(uuid.uuid4())
        num_imported = 0

        # Asumimos fila 1 como encabezado. Importamos desde fila 2.
        for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=min_required_cols, values_only=True), start=2):
            values = list(row)
            # Si la fila está completamente vacía, se salta
            if all(v is None for v in values):
                continue

            values_a_q = [
                str(values[0]) if len(values) > 0 and values[0] is not None else None,
                str(values[1]) if len(values) > 1 and values[1] is not None else None,
                str(values[2]) if len(values) > 2 and values[2] is not None else None,
                str(values[3]) if len(values) > 3 and values[3] is not None else None,
                str(values[4]) if len(values) > 4 and values[4] is not None else None,
                str(values[5]) if len(values) > 5 and values[5] is not None else None,
                str(values[6]) if len(values) > 6 and values[6] is not None else None,
                str(values[7]) if len(values) > 7 and values[7] is not None else None,
                str(values[8]) if len(values) > 8 and values[8] is not None else None,
                str(values[9]) if len(values) > 9 and values[9] is not None else None,
                str(values[10]) if len(values) > 10 and values[10] is not None else None,
                str(values[11]) if len(values) > 11 and values[11] is not None else None,
                str(values[12]) if len(values) > 12 and values[12] is not None else None,
                str(values[13]) if len(values) > 13 and values[13] is not None else None,
                str(values[14]) if len(values) > 14 and values[14] is not None else None,
                str(values[15]) if len(values) > 15 and values[15] is not None else None,
                str(values[16]) if len(values) > 16 and values[16] is not None else None,
            ]
            insert_raw_incident(document_id, idx, file_path, values_a_q)
            num_imported += 1

        logger.info("Importadas %s filas para document_id=%s", num_imported, document_id)
        return PrepareResponse(document_id=document_id, rows_imported=num_imported)
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Error al preparar hoja")
        raise HTTPException(status_code=500, detail=f"Error al preparar hoja: {exc}")


@app.post("/sheet/prepare-upload", response_model=PrepareResponse)
def prepare_sheet_upload(file: UploadFile = File(...)):
    """
    Recibe un archivo Excel subido, lo guarda y procesa usando la misma lógica que /sheet/prepare.
    Devuelve un document_id para identificar el dataset.
    """
    try:
        # Validar que sea un archivo Excel
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="El archivo debe ser un Excel (.xlsx o .xls)")
        
        # Generar document_id único
        document_id = str(uuid.uuid4())
        
        # Guardar archivo en UPLOADS_DIR
        upload_path = UPLOADS_DIR / f"{document_id}.xlsx"
        with open(upload_path, "wb") as buffer:
            content = file.file.read()
            buffer.write(content)
        
        logger.info("Archivo guardado en: %s", upload_path)
        
        # Usar la misma lógica que /sheet/prepare
        wb = load_workbook(upload_path)
        ws = wb.active
        
        # Validación: mínimo 17 columnas (A..Q)
        min_required_cols = 17
        if ws.max_column < min_required_cols:
            logger.error("El Excel no posee las columnas A-Q requeridas (encontradas: %s)", ws.max_column)
            raise HTTPException(status_code=400, detail="El Excel no posee las columnas A-Q requeridas")
        
        num_imported = 0
        
        # Asumimos fila 1 como encabezado. Importamos desde fila 2.
        for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=min_required_cols, values_only=True), start=2):
            values = list(row)
            # Si la fila está completamente vacía, se salta
            if all(v is None for v in values):
                continue
            
            values_a_q = [
                str(values[0]) if len(values) > 0 and values[0] is not None else None,
                str(values[1]) if len(values) > 1 and values[1] is not None else None,
                str(values[2]) if len(values) > 2 and values[2] is not None else None,
                str(values[3]) if len(values) > 3 and values[3] is not None else None,
                str(values[4]) if len(values) > 4 and values[4] is not None else None,
                str(values[5]) if len(values) > 5 and values[5] is not None else None,
                str(values[6]) if len(values) > 6 and values[6] is not None else None,
                str(values[7]) if len(values) > 7 and values[7] is not None else None,
                str(values[8]) if len(values) > 8 and values[8] is not None else None,
                str(values[9]) if len(values) > 9 and values[9] is not None else None,
                str(values[10]) if len(values) > 10 and values[10] is not None else None,
                str(values[11]) if len(values) > 11 and values[11] is not None else None,
                str(values[12]) if len(values) > 12 and values[12] is not None else None,
                str(values[13]) if len(values) > 13 and values[13] is not None else None,
                str(values[14]) if len(values) > 14 and values[14] is not None else None,
                str(values[15]) if len(values) > 15 and values[15] is not None else None,
                str(values[16]) if len(values) > 16 and values[16] is not None else None,
            ]
            insert_raw_incident(document_id, idx, str(upload_path), values_a_q)
            num_imported += 1
        
        logger.info("Importadas %s filas para document_id=%s", num_imported, document_id)
        return PrepareResponse(document_id=document_id, rows_imported=num_imported)
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Error al preparar hoja desde upload")
        raise HTTPException(status_code=500, detail=f"Error al preparar hoja desde upload: {exc}")


@app.get("/data/chunk/{document_id}", response_model=ChunkResponse)
def get_data_chunk(
    document_id: str = Path(..., description="Identificador del documento"),
    limit: int = Query(100, ge=1, le=1000, description="Cantidad máxima de filas a devolver"),
):
    try:
        rows = fetch_unclassified_chunk(document_id, limit)
        items = [
            {
                "id": r["id"],
                "row_index": r["row_index"],
                "col_a": r.get("col_a"),
                "col_b": r.get("col_b"),
                "col_c": r.get("col_c"),
                "col_d": r.get("col_d"),
                "col_e": r.get("col_e"),
                "col_f": r.get("col_f"),
                "col_g": r.get("col_g"),
                "col_h": r.get("col_h"),
                "col_i": r.get("col_i"),
                "col_j": r.get("col_j"),
                "col_k": r.get("col_k"),
                "col_l": r.get("col_l"),
                "col_m": r.get("col_m"),
                "col_n": r.get("col_n"),
                "col_o": r.get("col_o"),
                "col_p": r.get("col_p"),
                "col_q": r.get("col_q"),
            }
            for r in rows
        ]
        logger.info("Devueltos %s registros no clasificados para document_id=%s", len(items), document_id)
        return ChunkResponse(document_id=document_id, items=items)
    except Exception as exc:
        logger.exception("Error al obtener lote")
        raise HTTPException(status_code=500, detail=f"Error al obtener lote: {exc}")


@app.post("/data/save_classified_chunk", response_model=SaveClassifiedChunkResponse)
def save_classified_chunk(payload: SaveClassifiedChunkRequest):
    try:
        saved = insert_classified_items(
            payload.document_id,
            [item.model_dump() for item in payload.items],
        )
        logger.info("Guardados %s registros clasificados para document_id=%s", saved, payload.document_id)
        return SaveClassifiedChunkResponse(document_id=payload.document_id, saved=saved)
    except Exception as exc:
        logger.exception("Error al guardar clasificados")
        raise HTTPException(status_code=500, detail=f"Error al guardar clasificados: {exc}")


@app.post("/sheet/generate_final/{document_id}", response_model=GenerateFinalResponse)
def generate_final_sheet(document_id: str):
    try:
        # Traer todos los raw para el document_id
        raws = fetch_raws(document_id)
        if not raws:
            raise HTTPException(status_code=404, detail="No hay datos para el document_id indicado")

        # Mapear clasificados por raw_incident_id
        classified_by_raw = fetch_classified_map(document_id)

        # Construir nuevo Excel con encabezados A..Q y R..AB
        wb = Workbook()
        ws = wb.active
        ws.title = "final"

        headers_a_q = [
            "A",
            "B",
            "C",
            "D",
            "E",
            "F",
            "G",
            "H",
            "I",
            "J",
            "K",
            "L",
            "M",
            "N",
            "O",
            "P",
            "Q",
        ]
        headers_r_ab = [
            "R",
            "S",
            "T",
            "U",
            "V",
            "W",
            "X",
            "Y",
            "Z",
            "AA",
            "AB",
        ]
        ws.append(headers_a_q + headers_r_ab)

        # Color de relleno para columnas R-AB
        fill = PatternFill(fill_type="solid", start_color="FFB2A1C7", end_color="FFB2A1C7")

        # Escribir datos fila por fila
        for r in raws:
            a_q_values = [
                r.get("col_a"),
                r.get("col_b"),
                r.get("col_c"),
                r.get("col_d"),
                r.get("col_e"),
                r.get("col_f"),
                r.get("col_g"),
                r.get("col_h"),
                r.get("col_i"),
                r.get("col_j"),
                r.get("col_k"),
                r.get("col_l"),
                r.get("col_m"),
                r.get("col_n"),
                r.get("col_o"),
                r.get("col_p"),
                r.get("col_q"),
            ]
            c = classified_by_raw.get(r["id"])
            r_ab_values = [
                c.get("col_r") if c else None,
                c.get("col_s") if c else None,
                c.get("col_t") if c else None,
                c.get("col_u") if c else None,
                c.get("col_v") if c else None,
                c.get("col_w") if c else None,
                c.get("col_x") if c else None,
                c.get("col_y") if c else None,
                c.get("col_z") if c else None,
                c.get("col_aa") if c else None,
                c.get("col_ab") if c else None,
            ]
            ws.append(a_q_values + r_ab_values)

        # Aplicar color a todas las celdas de columnas R..AB (incluyendo encabezados)
        # Columnas R..AB corresponden a índices 18..28 (1-based)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=18, max_col=28):
            for cell in row:
                cell.fill = fill

        output_name = f"final_{document_id}.xlsx"
        output_path = FINAL_DIR / output_name
        wb.save(output_path)
        logger.info("Excel final generado: %s", output_path)
        return GenerateFinalResponse(document_id=document_id, file_path=str(output_path), rows=len(raws))
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Error al generar Excel final")
        raise HTTPException(status_code=500, detail=f"Error al generar Excel final: {exc}")


@app.get("/sheet/final/{document_id}")
def download_final_sheet(document_id: str):
    """
    Descarga el archivo Excel final generado para un document_id.
    Devuelve el archivo como FileResponse o 404 si no existe.
    """
    try:
        final_path = FINAL_DIR / f"final_{document_id}.xlsx"
        
        if not final_path.exists():
            raise HTTPException(status_code=404, detail="Archivo final no encontrado")
        
        logger.info("Descargando archivo final: %s", final_path)
        return FileResponse(
            path=str(final_path),
            filename=final_path.name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Error al descargar archivo final")
        raise HTTPException(status_code=500, detail=f"Error al descargar archivo final: {exc}")


